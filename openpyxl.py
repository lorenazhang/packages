import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Border, Side, Font

# Sample DataFrame with some URL columns
data = {
    "Name": ["Alice", "Bob", "Charlie"],
    "Age": [25, 30, 35],
    "City": ["New York", "Los Angeles", "Chicago"],
    "url_1": ["https://example.com/a/very/long/url/that/needs/wrapping", "https://example.com/b", "https://example.com/c"],
    "url_2": ["", "https://example.com/y", None],
}
df = pd.DataFrame(data)

# Save DataFrame to Excel (without index)
file_path = "output.xlsx"
df.to_excel(file_path, index=False, engine="openpyxl")

# Load the workbook and get the active sheet
wb = load_workbook(file_path)
ws = wb.active

# Freeze the top row (row 1)
ws.freeze_panes = "A2"

# Enable filter on the top row
ws.auto_filter.ref = ws.dimensions  # Apply filter to the entire range

# Define border style (thin border)
thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

# Define hyperlink font
hyperlink_font = Font(color="0000FF", underline="single")

# Adjust column width, wrap text, top align, apply borders and hyperlink URLs
for col_idx, col_cells in enumerate(ws.iter_cols(), 1):
    col_letter = get_column_letter(col_idx)
    header = col_cells[0].value
    max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col_cells) + 2
    ws.column_dimensions[col_letter].width = max(12, min(max_length, 50))  # Wider for links

    for row_idx, cell in enumerate(col_cells):
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        cell.border = thin_border

        # Skip header row
        if row_idx == 0:
            continue

        # If it's a URL column and a valid link
        if header and header.startswith("url_") and cell.value and isinstance(cell.value, str) and cell.value.startswith("http"):
            cell.hyperlink = cell.value
            cell.font = hyperlink_font  # Apply font manually instead of style
            cell.alignment = Alignment(wrap_text=True, vertical="top")

# Save the updated Excel file
wb.save(file_path)
print(f"Excel file '{file_path}' saved successfully with wrapped, clickable hyperlinks!")
